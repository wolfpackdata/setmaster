"""Set export: CSV / XLSX / Markdown generators + endpoint (set-export.md).

Content-level assertions: metadata header, grid columns in grid order, derived
Out side, m:ss timing text, Key display variants (flats vs camelot),
`---` placeholders, RFC 4180 quoting, XLSX native fills/borders, and the
Content-Disposition filename (§5).
"""
from __future__ import annotations

import csv
import io
import uuid

from openpyxl import load_workbook

from test_app_helpers import app_client

RID0 = "row-aaaa"
RID1 = "row-bbbb"


def _rows() -> list[dict]:
    def base(**o):
        r = {
            "id": uuid.uuid4().hex, "bpm": "", "key": "", "in_name": "",
            "in_delta": "---", "m_num": "---", "t_num": "---", "a_num": "---",
            "lows": "", "level": "", "swap_lows": "---", "i_like": "⚠️",
            "notes": "", "start": "", "transition": "",
        }
        r.update(o)
        return r

    return [
        base(
            id=RID0, bpm="124", key="Gbm", in_name="Track One", in_delta="+1",
            m_num="#1", lows="cut", level="HOT", i_like="🚀", notes="plain",
            start="0:00", transition="1:00",
        ),
        base(
            id=RID1, bpm="126", key="gbm", in_name="Track, Two\nline",
            in_delta="-0.5", m_num="#2", t_num="#3", swap_lows="#4",
            notes="note, with comma\nand newline | pipe",
            start="1:00", transition="2:30",
        ),
    ]


def _make_set(c, name="Kimma Bryan", *, fills=None, boxes=None) -> str:
    sid = c.post("/api/sets", json={"name": name}).json()["id"]
    assert c.put(f"/api/sets/{sid}/rows", json=_rows()).status_code == 200
    if fills or boxes:
        r = c.put(
            f"/api/sets/{sid}/formatting",
            json={"fills": fills or [], "boxes": boxes or []},
        )
        assert r.status_code == 200
    return sid


EXPECTED_HEADER = [
    "BPM", "Key", "Out Track Name", "Out Δ", "T #", "A #", "In Track Name",
    "In Δ", "M #", "Lows", "Level", "Swap Lows", "I like", "FX & Mix Notes",
    # #104 — the #72 grid vocabulary, prefixed with the group. Bare `M #`/`T #`
    # would duplicate the cue headers five columns to the left.
    "Out M #", "Out T #",
]


def test_export_header_has_no_duplicate_labels():
    """#104 — a flat file has no super-header to disambiguate two `M #` columns.

    This is the constraint that decided the rename's form, so it is asserted
    rather than left to a comment: pandas silently renames a duplicate to
    `M #.1`, and a reader cannot tell the cue column from the timing column.
    """
    assert len(EXPECTED_HEADER) == len(set(EXPECTED_HEADER))
    # The cue columns keep their SM2-verbatim labels; only the timing pair moved.
    assert "T #" in EXPECTED_HEADER and "M #" in EXPECTED_HEADER
    assert "Start" not in EXPECTED_HEADER and "Transition" not in EXPECTED_HEADER


# --- CSV --------------------------------------------------------------------


def test_csv_content_and_quoting(tmp_path):
    with app_client(tmp_path) as c:
        sid = _make_set(c)
        r = c.post(
            f"/api/sets/{sid}/export",
            json={"format": "csv", "key_display_as": "flats"},
        )
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("text/csv")
        # BOM so Excel opens emoji / 'Gbm' correctly (§6.1).
        assert r.content.startswith(b"\xef\xbb\xbf")

        text = r.content.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        # §3.1: 4 comment rows, blank row, header at row 6 (index 5).
        assert rows[0] == ["# Set: Kimma Bryan"]
        assert rows[1][0].startswith("# Exported: ")
        assert rows[2] == ["# Tracks: 2"]
        assert rows[3] == ["# Total mix length: 2.5"]
        assert rows[4] == []
        assert rows[5] == EXPECTED_HEADER

        d0, d1 = rows[6], rows[7]
        # Row 0: derived Out side is blank/'---'; m:ss verbatim; placeholders.
        assert d0[0] == "124" and d0[1] == "Gbm"
        assert d0[2] == "" and d0[3] == "---"  # Out Track Name / Out Δ (first row)
        assert d0[4] == "---" and d0[5] == "---"  # T # / A #
        assert d0[6] == "Track One" and d0[7] == "+1"
        assert d0[12] == "🚀"
        assert d0[14] == "0:00" and d0[15] == "1:00"
        # Row 1: Out side mirrors row 0's In side; lowercase key canonicalizes.
        assert d1[1] == "Gbm"  # 'gbm' typed -> canonical display
        assert d1[2] == "Track One" and d1[3] == "+1"
        assert d1[6] == "Track, Two\nline"  # comma + newline preserved in-cell
        assert d1[11] == "#4"  # Swap Lows
        assert d1[13] == "note, with comma\nand newline | pipe"


def test_csv_key_display_camelot(tmp_path):
    with app_client(tmp_path) as c:
        sid = _make_set(c)
        r = c.post(
            f"/api/sets/{sid}/export",
            json={"format": "csv", "key_display_as": "camelot"},
        )
        rows = list(csv.reader(io.StringIO(r.content.decode("utf-8-sig"))))
        # Gbm -> 11A in Camelot (both rows).
        assert rows[6][1] == "11A"
        assert rows[7][1] == "11A"


# --- Markdown ---------------------------------------------------------------


def test_markdown_table_and_escaping(tmp_path):
    with app_client(tmp_path) as c:
        sid = _make_set(c)
        r = c.post(
            f"/api/sets/{sid}/export",
            json={"format": "markdown", "key_display_as": "flats"},
        )
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("text/markdown")
        text = r.content.decode("utf-8")
        lines = text.splitlines()
        assert lines[0] == "# Kimma Bryan"
        assert lines[2].startswith("**Exported:** ")
        assert "**Tracks:** 2" in lines[2]
        assert "**Total mix length:** 2.5" in lines[2]
        assert lines[4] == "| " + " | ".join(EXPECTED_HEADER) + " |"
        assert lines[5] == "| " + " | ".join("---" for _ in EXPECTED_HEADER) + " |"
        # Pipe escaped, in-cell newline -> <br> (§6.3).
        assert "note, with comma<br>and newline \\| pipe" in text
        assert "Track, Two<br>line" in text


# --- XLSX -------------------------------------------------------------------


def test_xlsx_structure_types_and_formatting(tmp_path):
    with app_client(tmp_path) as c:
        sid = _make_set(
            c,
            fills=[{"row_id": RID0, "col": "bpm", "color": "red"}],
            boxes=[{"row_ids": [RID0, RID1], "cols": ["lows", "level"]}],
        )
        r = c.post(
            f"/api/sets/{sid}/export",
            json={"format": "xlsx", "key_display_as": "flats"},
        )
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]

        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.title == "Kimma Bryan"
        # §3.1 metadata block.
        assert ws["A1"].value == "Set" and ws["B1"].value == "Kimma Bryan"
        assert ws["A3"].value == "Tracks" and ws["B3"].value == 2
        assert ws["B4"].value == "2.5"
        # Header row 6, bold; frozen below.
        assert [ws.cell(row=6, column=i + 1).value for i in range(16)] == EXPECTED_HEADER
        assert ws.cell(row=6, column=1).font.bold is True
        assert ws.freeze_panes == "A7"

        # BPM as a number; Start/Transition as text (m:ss).
        assert ws.cell(row=7, column=1).value == 124
        assert ws.cell(row=7, column=15).value == "0:00"
        assert isinstance(ws.cell(row=7, column=15).value, str)
        # Derived Out side + placeholder.
        assert ws.cell(row=7, column=3).value in (None, "")  # Out Track Name (row 0)
        assert ws.cell(row=7, column=4).value == "---"  # Out Δ placeholder
        assert ws.cell(row=8, column=3).value == "Track One"

        # RED fill on BPM of row 0 (native fill).
        fill = ws.cell(row=7, column=1).fill
        assert fill.patternType == "solid"
        assert (fill.fgColor.rgb or "").endswith("FFC7CE")
        # Box border around the lows/level rectangle (cols 10-11, rows 7-8).
        top_left = ws.cell(row=7, column=10).border
        assert top_left.top.style == "medium" and top_left.left.style == "medium"
        bottom_right = ws.cell(row=8, column=11).border
        assert bottom_right.bottom.style == "medium" and bottom_right.right.style == "medium"


# --- Filename (§5) + validation ---------------------------------------------


def test_export_filename_default_and_remembered(tmp_path):
    with app_client(tmp_path) as c:
        sid = _make_set(c, name="Kimma Bryan")
        # Default: <slug>_<date>.<ext>
        r = c.post(
            f"/api/sets/{sid}/export",
            json={"format": "csv", "key_display_as": "flats"},
        )
        disp = r.headers["content-disposition"]
        assert 'filename="kimma-bryan_' in disp and ".csv" in disp

        # Remembered name, extension swapped to the chosen format.
        assert c.patch(
            f"/api/sets/{sid}/export-filename", json={"filename": "my-set.xlsx"}
        ).status_code == 200
        r = c.post(
            f"/api/sets/{sid}/export",
            json={"format": "markdown", "key_display_as": "flats"},
        )
        assert 'filename="my-set.md"' in r.headers["content-disposition"]


def test_export_rejects_bad_params(tmp_path):
    with app_client(tmp_path) as c:
        sid = _make_set(c)
        assert c.post(
            f"/api/sets/{sid}/export",
            json={"format": "pdf", "key_display_as": "flats"},
        ).status_code == 400
        assert c.post(
            f"/api/sets/{sid}/export",
            json={"format": "csv", "key_display_as": "bogus"},
        ).status_code == 400
        assert c.post(
            "/api/sets/nope/export",
            json={"format": "csv", "key_display_as": "flats"},
        ).status_code == 404
