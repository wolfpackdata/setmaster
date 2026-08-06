"""Server-side set export: CSV / XLSX / Markdown (planning/02-features/set-export.md).

Renders the set EXACTLY AS DISPLAYED in the S2 grid (data-model §4.2; ui §5.2):

- the derived Out side (out_name / out_delta mirror the previous row's In side),
- m:ss timing text (Start / Transition) shown verbatim,
- Key rendered in the requested notation (Key Display As) when the typed text
  parses as a key in any of the 4 notations; free text passes through untouched,
- enum cells show the on-screen ``---`` placeholder when blank (§3.2),
- the four-field metadata header (§3.1), then the 16 grid columns in grid order.

openpyxl (already a backend dependency) generates the XLSX; CSV and Markdown are
hand-rolled. Everything is pure over (name, rows, formatting, key_display_as) so
the router just fetches the set and streams the bytes back.
"""
from __future__ import annotations

import csv
import datetime
import io
import math
import os
import re
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .keys import KEY_TABLE, to_notation
from .util import now_iso

# ---------------------------------------------------------------------------
# Format metadata
# ---------------------------------------------------------------------------

MEDIA_TYPE = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "markdown": "text/markdown; charset=utf-8",
}

_EXT = {"csv": "csv", "xlsx": "xlsx", "markdown": "md"}
# Extensions we recognise on a remembered filename so we can swap in the one
# matching the chosen format (a user who exported XLSX then picks CSV keeps the
# same base name, .csv extension).
_KNOWN_EXT = {".csv", ".xlsx", ".md", ".markdown", ".txt"}


# ---------------------------------------------------------------------------
# Column model — grid order + exact labels (§3.2)
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ExportCol:
    key: str  # row field, or 'out_name' / 'out_delta' (derived)
    label: str
    placeholder: bool  # blank cell renders the on-screen '---' placeholder
    numeric: bool  # XLSX: write as a number when the display text parses


# Column order and labels per set-export.md §3.2. The flat export has no
# out/in group headers, so the derived-name/delta columns carry the fuller
# "Out Track Name" / "Out Δ" / "In Track Name" / "In Δ" labels named there.
EXPORT_COLS: tuple[ExportCol, ...] = (
    ExportCol("bpm", "BPM", False, True),
    ExportCol("key", "Key", False, False),
    ExportCol("out_name", "Out Track Name", False, False),
    ExportCol("out_delta", "Out Δ", True, False),
    ExportCol("t_num", "T #", True, False),
    ExportCol("a_num", "A #", True, False),
    ExportCol("in_name", "In Track Name", False, False),
    ExportCol("in_delta", "In Δ", True, False),
    ExportCol("m_num", "M #", True, False),
    ExportCol("lows", "Lows", True, False),
    ExportCol("level", "Level", True, False),
    ExportCol("swap_lows", "Swap Lows", True, False),
    ExportCol("i_like", "I like", False, False),
    ExportCol("notes", "FX & Mix Notes", False, False),
    # #104 (D-049 resolved 2026-07-29) — the flat export adopts the #72 grid
    # vocabulary, but PREFIXED with the group rather than bare. Bare `M #` / `T #`
    # would collide with the cue columns above: two `M #` and two `T #` in one
    # flat file, where pandas silently renames the second and a reader cannot
    # tell which is which. The grid is only unambiguous because the OUT TRACK
    # TIMING super-header sits above these two; a flat file has no such row, so
    # the super-header is folded into the label instead. Closest possible
    # translation of what is on screen.
    ExportCol("start", "Out M #", False, False),
    ExportCol("transition", "Out T #", False, False),
)

_PLACEHOLDER = "---"


# ---------------------------------------------------------------------------
# Key notation (case-insensitive, mirrors the grid's canonicalizeKey + formatKey)
# ---------------------------------------------------------------------------

# The set-row Key is manually typed, so any of the 4 notations may appear in any
# case (`gbm`, `F#M`, `11a`). Build a case-insensitive map to the canonical
# flats key; app/keys.py's from_any_notation is case-sensitive, so this local
# lookup matches what the S2 grid renders.
_CANON_CI: dict[str, str] = {}
for _row in KEY_TABLE:
    _flats = _row[0]
    for _val in _row:
        _CANON_CI.setdefault(_val.lower(), _flats)


def _render_key(raw: str, key_display_as: str) -> str:
    text = raw.strip()
    if not text:
        return ""
    canonical = _CANON_CI.get(text.lower())
    if canonical is None:
        return text  # free text — passes through untouched (§4.2)
    return to_notation(canonical, key_display_as) or text


# ---------------------------------------------------------------------------
# Timing (ported verbatim from the S2 model — model.ts parseMss/computeTiming)
# ---------------------------------------------------------------------------

_MSS_RE = re.compile(r"^\s*(\d{1,3}):([0-5]\d)\s*$")


def _round2(n: float) -> float:
    # Half-up to 2 decimals (JS Math.round semantics; timing is non-negative).
    return math.floor(n * 100 + 0.5) / 100


def _parse_mss(text: str) -> int | None:
    m = _MSS_RE.match(text or "")
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2))


def _row_minutes(start: str, transition: str) -> float | None:
    s = _parse_mss(start)
    t = _parse_mss(transition)
    if s is None or t is None:
        return None
    diff = t - s
    if diff < 0:
        return None
    return _round2(diff / 60)


def _fmt_minutes(n: float) -> str:
    r = _round2(n)
    if r == int(r):
        return str(int(r))
    return f"{r:.2f}".rstrip("0").rstrip(".")


def _mix_length_text(rows: list[dict]) -> str:
    """Stats 'Mix Length' = max cumulative minutes; blank when no timing (§3.1)."""
    running = 0.0
    seen = False
    for r in rows:
        m = _row_minutes(r.get("start", ""), r.get("transition", ""))
        if m is not None:
            running = _round2(running + m)
            seen = True
    return _fmt_minutes(running) if seen else ""


def _track_count(rows: list[dict]) -> int:
    # SM2 '# Tracks': rows with a non-empty In Track name (matches the on-screen
    # Stats value and the /api/sets track_count; empty scaffold rows don't count).
    return sum(1 for r in rows if (r.get("in_name") or "").strip() != "")


# ---------------------------------------------------------------------------
# Cell rendering
# ---------------------------------------------------------------------------


def _derive_outs(rows: list[dict]) -> list[tuple[str, str]]:
    outs: list[tuple[str, str]] = []
    for i, _ in enumerate(rows):
        if i == 0:
            outs.append(("", ""))
        else:
            prev = rows[i - 1]
            outs.append(((prev.get("in_name") or "").strip(), (prev.get("in_delta") or "").strip()))
    return outs


def _cell_text(
    rows: list[dict], outs: list[tuple[str, str]], i: int, col: ExportCol, key_display_as: str
) -> str:
    if col.key == "key":
        return _render_key(rows[i].get("key") or "", key_display_as)
    if col.key == "out_name":
        raw = outs[i][0]
    elif col.key == "out_delta":
        raw = outs[i][1]
    else:
        raw = rows[i].get(col.key) or ""
    if col.placeholder and (raw.strip() == "" or raw.strip() == _PLACEHOLDER):
        return _PLACEHOLDER
    return raw


def _build_matrix(rows: list[dict], key_display_as: str) -> list[list[str]]:
    outs = _derive_outs(rows)
    return [
        [_cell_text(rows, outs, i, col, key_display_as) for col in EXPORT_COLS]
        for i in range(len(rows))
    ]


# ---------------------------------------------------------------------------
# Filename (§5)
# ---------------------------------------------------------------------------


def slugify(name: str) -> str:
    """§5: lowercase, spaces -> '-', strip filesystem-illegal characters."""
    s = name.strip().lower().replace(" ", "-")
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or "set"


def export_filename_for(export_filename: str | None, set_name: str, fmt: str) -> str:
    """Content-Disposition filename: the set's remembered name (extension swapped
    to the chosen format) or the §5 default ``<slug>_<YYYY-MM-DD>.<ext>``."""
    ext = _EXT[fmt]
    if export_filename and export_filename.strip():
        base = export_filename.strip()
        stem, dot = os.path.splitext(base)
        if dot.lower() in _KNOWN_EXT:
            base = stem
    else:
        today = datetime.datetime.now(datetime.timezone.utc).date().isoformat()
        base = f"{slugify(set_name)}_{today}"
    return f"{base}.{ext}"


# ---------------------------------------------------------------------------
# Generators
# ---------------------------------------------------------------------------


def _metadata(name: str, rows: list[dict]) -> tuple[str, str, int, str]:
    return name, now_iso(), _track_count(rows), _mix_length_text(rows)


def _render_csv(name: str, rows: list[dict], key_display_as: str) -> bytes:
    n, exported, tracks, mix = _metadata(name, rows)
    matrix = _build_matrix(rows, key_display_as)
    buf = io.StringIO()
    w = csv.writer(buf)  # RFC 4180 minimal quoting (commas/quotes/newlines)
    # §3.1: four comment rows, a blank row, then the header row (row 6).
    w.writerow([f"# Set: {n}"])
    w.writerow([f"# Exported: {exported}"])
    w.writerow([f"# Tracks: {tracks}"])
    w.writerow([f"# Total mix length: {mix}"])
    w.writerow([])
    w.writerow([c.label for c in EXPORT_COLS])
    for row in matrix:
        w.writerow(row)
    # utf-8-sig: BOM so Excel opens emoji and 'Gbm'-style text correctly (§6.1).
    return buf.getvalue().encode("utf-8-sig")


def _md_escape(text: str) -> str:
    # §6.3: escape pipes; in-cell newlines become <br>.
    return text.replace("\\", "\\\\").replace("|", "\\|").replace("\r\n", "\n").replace(
        "\n", "<br>"
    )


def _render_markdown(name: str, rows: list[dict], key_display_as: str) -> bytes:
    n, exported, tracks, mix = _metadata(name, rows)
    matrix = _build_matrix(rows, key_display_as)
    header = [c.label for c in EXPORT_COLS]
    lines = [
        f"# {n}",
        "",
        f"**Exported:** {exported} · **Tracks:** {tracks} · "
        f"**Total mix length:** {mix}",
        "",
        "| " + " | ".join(_md_escape(h) for h in header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in matrix:
        lines.append("| " + " | ".join(_md_escape(c) for c in row) + " |")
    return ("\n".join(lines) + "\n").encode("utf-8")


_XLSX_RED = PatternFill("solid", fgColor="FFC7CE")
_XLSX_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_XLSX_BOX_SIDE = Side(style="medium", color="000000")

# Sensible widths (§6.2): names and notes wide, cue/enum columns narrow.
_XLSX_WIDTHS = {
    "bpm": 8, "key": 8, "out_name": 30, "out_delta": 8, "t_num": 6, "a_num": 6,
    "in_name": 30, "in_delta": 8, "m_num": 6, "lows": 12, "level": 12,
    "swap_lows": 12, "i_like": 8, "notes": 46, "start": 10, "transition": 12,
}

_HEADER_ROW = 6
_FIRST_DATA_ROW = 7


def _sheet_name(name: str) -> str:
    s = re.sub(r"[\[\]:*?/\\]", " ", name).strip()
    return (s or "Set")[:31]


def _render_xlsx(name: str, rows: list[dict], formatting: dict, key_display_as: str) -> bytes:
    n, exported, tracks, mix = _metadata(name, rows)
    matrix = _build_matrix(rows, key_display_as)

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_name(name)

    # §3.1 layout: rows 1-4 label/value pairs, row 5 blank, header at row 6.
    for r, (label, value) in enumerate(
        (("Set", n), ("Exported", exported), ("Tracks", tracks), ("Total mix length", mix)),
        start=1,
    ):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)

    for ci, col in enumerate(EXPORT_COLS, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=ci, value=col.label)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(ci)].width = _XLSX_WIDTHS[col.key]

    for ri, mrow in enumerate(matrix):
        excel_row = _FIRST_DATA_ROW + ri
        for ci, col in enumerate(EXPORT_COLS, start=1):
            val = mrow[ci - 1]
            if col.numeric and val not in ("", _PLACEHOLDER):
                try:
                    ws.cell(row=excel_row, column=ci, value=float(val))
                    continue
                except ValueError:
                    pass
            # Start/Transition stay text so Excel keeps 'm:ss' (§6.2 — avoids the
            # 12:00 AM time misread); everything else is text too.
            ws.cell(row=excel_row, column=ci, value=val)

    # Bold header frozen at the top; autofilter over the header + data.
    ws.freeze_panes = f"A{_FIRST_DATA_ROW}"
    last_row = max(_HEADER_ROW + len(matrix), _HEADER_ROW)
    ws.auto_filter.ref = f"A{_HEADER_ROW}:{get_column_letter(len(EXPORT_COLS))}{last_row}"

    _apply_xlsx_formatting(ws, rows, formatting)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _apply_xlsx_formatting(ws, rows: list[dict], formatting: dict) -> None:
    """RED/YELLOW fills + Box borders as native cell styling (§3.2 — XLSX only)."""
    col_index = {c.key: i + 1 for i, c in enumerate(EXPORT_COLS)}
    row_index = {r.get("id"): i for i, r in enumerate(rows)}

    for f in formatting.get("fills", []):
        ri = row_index.get(f.get("row_id"))
        ci = col_index.get(f.get("col"))
        if ri is None or ci is None:
            continue
        ws.cell(row=_FIRST_DATA_ROW + ri, column=ci).fill = (
            _XLSX_RED if f.get("color") == "red" else _XLSX_YELLOW
        )

    for b in formatting.get("boxes", []):
        r_idxs = [row_index[x] for x in b.get("row_ids", []) if x in row_index]
        c_idxs = [col_index[x] for x in b.get("cols", []) if x in col_index]
        if not r_idxs or not c_idxs:
            continue
        r0, r1 = min(r_idxs), max(r_idxs)
        c0, c1 = min(c_idxs), max(c_idxs)
        for rr in range(r0, r1 + 1):
            for cc in range(c0, c1 + 1):
                top, bottom = rr == r0, rr == r1
                left, right = cc == c0, cc == c1
                if not (top or bottom or left or right):
                    continue  # interior cell — leave its styling untouched
                cell = ws.cell(row=_FIRST_DATA_ROW + rr, column=cc)
                cur = cell.border
                cell.border = Border(
                    top=_XLSX_BOX_SIDE if top else cur.top,
                    bottom=_XLSX_BOX_SIDE if bottom else cur.bottom,
                    left=_XLSX_BOX_SIDE if left else cur.left,
                    right=_XLSX_BOX_SIDE if right else cur.right,
                )


def render_export(
    fmt: str, name: str, rows: list[dict], formatting: dict, key_display_as: str
) -> bytes:
    """Generate the export bytes for one of csv / xlsx / markdown."""
    if fmt == "csv":
        return _render_csv(name, rows, key_display_as)
    if fmt == "xlsx":
        return _render_xlsx(name, rows, formatting, key_display_as)
    if fmt == "markdown":
        return _render_markdown(name, rows, key_display_as)
    raise ValueError(f"unknown export format: {fmt!r}")
